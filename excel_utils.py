import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Function to save descriptions to an Excel file
def save_to_excel(descriptions, ai_descriptions_file):
    try:
        # Convert the descriptions list to a DataFrame
        df = pd.DataFrame(descriptions)
        # Save the DataFrame to an Excel file
        df.to_excel(ai_descriptions_file, index=False)
    except Exception as e:
        # Print an error message if saving to Excel fails
        print(f"Error saving to Excel: {e}")
        raise

# Function to adjust the formatting of the Excel file
def adjust_excel_formatting(ai_descriptions_file, max_column_width=50):
    try:
        # Load the Excel workbook and select the active worksheet
        wb = load_workbook(ai_descriptions_file)
        ws = wb.active

        # Adjust the width of each column
        for col in ws.columns:
            column = col[0].column_letter
            ws.column_dimensions[column].width = max_column_width
            for cell in col:
                # Set cell alignment to wrap text and align vertically to the top
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Reset the height of the first row
        ws.row_dimensions[1].height = None

        # Adjust the height of each row and set cell alignment
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row[0].row].height = 200

        # Save the changes to the Excel file
        wb.save(ai_descriptions_file)
        print(f'Saved all AI descriptions to {ai_descriptions_file}')
    except Exception as e:
        # Print an error message if adjusting Excel formatting fails
        print(f"Error adjusting Excel formatting: {e}")
        raise